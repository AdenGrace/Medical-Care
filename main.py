from tkinter import *

import pandas as pd
import xlsxwriter as xlsxwriter
from PIL import Image, ImageTk
from tkinter import messagebox
from tkinter import scrolledtext
import xlsxwriter
import openpyxl,xlrd
import pathlib
from openpyxl import Workbook



window=Tk()
window.title("log in page")
window.geometry("600x400")
window.configure(bg="#E3E3E3")




file = pathlib.Path("Diagnose2.xlsx")
if file.exists():
        pass
else:
        file=Workbook()
        sheet=file.active
        sheet["A1"] = "Name"
        sheet["B1"] = "ID"
        sheet["C1"] = "Gender"
        sheet["D1"] = "Age"
        sheet["E1"] = "WBC"
        sheet["F1"] = "Neut"
        sheet["G1"] = "Lymph"
        sheet["H1"] = "RBC"
        sheet["I1"] = "HCT"
        sheet["J1"] = "Urea"
        sheet["K1"] = "Hb"
        sheet["L1"] = "Creatinine"
        sheet["M1"] = "Iron"
        sheet["N1"] = "HDL"
        sheet["O1"] = "AlkaLine"

        file.save("Diagnose2.xlsx")

def thirdPage():


        window.geometry("750x750")
        window.title("Patient's Details:")
        f3 = Frame()
        f3['bg'] = "#e3e3e3"
        f3.place(x=0, y=0, width=750, height=750)
        Nl1 = Label(f3, text="Patient Details", fg="black", bg="#e3e3e3", font=("times", 23, "bold"))
        Nl1.place(x=150, y=30)

        Nl2 = Label(f3, text="Name :", fg="black", bg="#e3e3e3", font=("times", 15))
        Nl2.place(x=50, y=90)
        NUserN = Entry(f3, borderwidth=3)
        NUserN.place(x=150, y=95)

        Nl3 = Label(f3, text="Id :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl3.place(x=50, y=140)
        NIdN = Entry(f3, borderwidth=3)
        NIdN.place(x=150, y=145)


        Nl4 = Label(f3, text="Gender :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl4.place(x=50, y=190)
        i = StringVar()
        R1 = Radiobutton(f3, text="Male",value="1", fg="black", bg="#e3e3e3",variable=i)
        R1.place(x=150, y=195)
        R1 = Radiobutton(f3, text="Female",value="2", fg="black", bg="#e3e3e3",variable=i)
        R1.place(x=220, y=195)

        def Gender():
                if (i.get()) == "1":
                        return "Male"
                elif (i.get()) == "2":
                        return "Female"
                else:
                        return " "

        Nl5 = Label(f3, text="age :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl5.place(x=50, y=240)
        Nl5N = Entry(f3, borderwidth=3)
        Nl5N.place(x=150, y=245)


        Nl6 = Label(f3, text="WBC :", fg="black", bg="#e3e3e3", font=("times", 15))
        Nl6.place(x=50, y=290)
        Nl6N = Entry(f3, borderwidth=3)
        Nl6N.place(x=150, y=295)

        Nl7 = Label(f3, text="Neut :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl7.place(x=50, y=340)
        Nl7N = Entry(f3, borderwidth=3)
        Nl7N.place(x=150, y=345)

        Nl8 = Label(f3, text="Lymph :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl8.place(x=50, y=390)
        Nl8N = Entry(f3, borderwidth=3)
        Nl8N.place(x=150, y=395)

        Nl9 = Label(f3, text="RBC :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl9.place(x=50, y=440)
        Nl9N = Entry(f3, borderwidth=3)
        Nl9N.place(x=150, y=445)

        Nl10 = Label(f3, text="HCT :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl10.place(x=50, y=490)
        Nl10N = Entry(f3, borderwidth=3)
        Nl10N.place(x=150, y=495)

        Nl11 = Label(f3, text="Urea :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl11.place(x=50, y=540)
        Nl11N = Entry(f3, borderwidth=3)
        Nl11N.place(x=150, y=545)

        Nl12 = Label(f3, text="Hb :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl12.place(x=50, y=590)
        Nl12N = Entry(f3, borderwidth=3)
        Nl12N.place(x=150, y=595)

        Nl13 = Label(f3, text="Creatine :", fg="black", bg="#e3e3e3", font=("times", 15))
        Nl13.place(x=50, y=640)
        Nl13N = Entry(f3, borderwidth=3)
        Nl13N.place(x=150, y=645)

        Nl14 = Label(f3, text="Iron :",  fg="black", bg="#e3e3e3", font=("times", 15))
        Nl14.place(x=320, y=90)
        Nl14N = Entry(f3, borderwidth=3)
        Nl14N.place(x=550, y=95)

        Nl15 = Label(f3, text="HDL :", fg="black", bg="#e3e3e3", font=("times", 15))
        Nl15.place(x=320, y=140)
        Nl15N = Entry(f3, borderwidth=3)
        Nl15N.place(x=550, y=145)

        Nl16 = Label(f3, text="Alkaline Phosphatase :", fg="black", bg="#e3e3e3", font=("times", 15))
        Nl16.place(x=320, y=190)
        Nl16N = Entry(f3, borderwidth=3)
        Nl16N.place(x=550, y=195)

        CheckVar1 = IntVar()
        CheckVar2 = IntVar()
        CheckVar3 = IntVar()
        CheckVar4 = IntVar()
        C1 = Checkbutton(f3, text="Are you Smoking?", fg="black", bg="#e3e3e3", variable=CheckVar1, onvalue=1, offvalue=0, height=5,width=20)
        C2 = Checkbutton(f3, text="Are you Pregnant?", fg="black", bg="#e3e3e3", variable=CheckVar2,onvalue=1, offvalue=0, height=5,width=20)
        C3 = Checkbutton(f3, text="Are you Asian?", fg="black", bg="#e3e3e3", variable=CheckVar3, onvalue=1, offvalue=0, height=5,width=20)
        C4 = Checkbutton(f3, text="Are you Ethiopian?", fg="black", bg="#e3e3e3", variable=CheckVar4, onvalue=1, offvalue=0, height=5,width=20)

        C1.place(x=320,y=240)
        C2.place(x=320,y=290)
        C3.place(x=310,y=340)
        C4.place(x=320,y=390)


        def EndPage():
                f4 = Frame()
                window.title("Patient Diagnose")
                f4['bg'] = "#E3E3E3"
                f4.place(x=0, y=0, width=750, height=750)
                tf=open('Diagnose.txt','r+')

                data = tf.read()
                txtarea = scrolledtext.ScrolledText(f4,wrap=WORD,width=600,height=600)
                txtarea.place(x=45, y=15, width=650, height=650)
                txtarea['font']="times 14"
                txtarea.insert(END,data)
                tf.close()
                txtarea.configure(state='disabled')
                exitButton=Button(f4, text="Close", bg="#91c6cf", fg="white",command=secondPage, font=("times", 15, "bold"))
                exitButton.place(x=650, y=690)

        def check():
                if (any(c.isalpha() for c in Nl5N.get())==True or any(c.isalpha() for c in Nl6N.get())==True
                   or any(c.isalpha() for c in Nl7N.get())==True or any(c.isalpha() for c in Nl8N.get())==True or any(c.isalpha() for c in Nl9N.get())==True
                   or any(c.isalpha() for c in Nl10N.get()) == True or any(c.isalpha() for c in Nl11N.get())==True or any(c.isalpha() for c in Nl12N.get())==True
                   or any(c.isalpha() for c in Nl13N.get()) == True or any(c.isalpha() for c in Nl14N.get())==True or any(c.isalpha() for c in Nl15N.get())==True
                   or any(c.isalpha() for c in Nl16N.get())==True):
                        return False

                return True

        def information():
                if len(Nl6N.get())!=0 and len(Nl7N.get())!=0 and len(Nl8N.get())!=0 and len(Nl9N.get())!=0 and len(Nl10N.get())!=0 and len(Nl11N.get())!=0 and\
                        len(Nl12N.get())!=0 and len(Nl13N.get())!=0 and len(Nl14N.get())!=0 and Gender()!=" " :
                        if check()==False:
                                messagebox.showinfo("Error","Details must be numbers")
                        else:
                                txtfile()
                else:
                        messagebox.showinfo("Error", "write all the values")

        Nb1 = Button(f3, text="Diagnose", bg="#91c6cf", fg="white", command=information, font=("times", 15, "bold"))
        Nb1.place(x=600, y=700)

 # *****************************************************
        def HCT(fh, hct,smoke):
                file = openpyxl.load_workbook("Diagnose2.xlsx")
                sheet = file.active
                if (Gender()=="female" and hct<33)or (Gender()=="male" and hct<37):
                        sheet.cell(column=9, row=sheet.max_row, value="Most often indicate bleeding or anemia.")
                        sheet.cell(column=9, row=sheet.max_row, value="therapy: in bleeding case To be evacuated urgently to the hospital in anemia caseTwo 10 mg B12 pills a day for a month")
                        fh.write('\nMost often indicate bleeding or anemia.')
                        fh.write('\ntherapy: in bleeding case To be evacuated urgently to the hospital in anemia'
                                 ' caseTwo 10 mg B12 pills a day for a month')
                elif (Gender()=="female" and hct>47 and smoke==1)or \
                        (Gender()=="male" and hct>54 and smoke==1):
                        fh.write('\nCommon in smokers.')
                        fh.write('\ntherapy: stop smoking')
                        sheet.cell(column=9, row=sheet.max_row, value="Common in smokers")
                        sheet.cell(column=9, row=sheet.max_row, value="therapy: stop smoking")

                else:
                        fh.write('\nIts ok')
                        sheet.cell(column=9, row=sheet.max_row, value="\Its Ok")
                file.save("Diagnose2.xlsx")
# *****************************************************
        def Hb(fh, hb, age):
                file = openpyxl.load_workbook("Diagnose2.xlsx")
                sheet = file.active
                if (Gender()=="female" and age>=18 and hb>=12 and hb<=16 )or(Gender()=="male" and age>=18 and hb>=12 and hb<=18) or(age>=0 and age<=17 and hb>=11.5 and hb<=15.5):
                        sheet.cell(column=11, row=sheet.max_row, value="\Its Ok")
                        fh.write('\nHb: \nIts ok')
                elif (Gender()=="female" and age>=18 and hb<12 )or(Gender()=="male" and age>=18 and hb<12) or(age>=0 and age<=17 and hb<11.5 ):
                        fh.write(
                                '\nHb: \nIndicates anemia. This can be due to hematologic disorder, iron deficiency and bleeding.')
                        fh.write(
                                '\nHb: \ntherapy: in case of anemia Two 10 mg B12 pills a day for a month, in case of hematologic'
                                ' disorder An injection of a hormone to encourage red blood cell production, in case of iron deficiency'
                                ' Two 10 mg pills of B12 a day for a month, in case of bleeding To be evacuated urgently to the hospital')
                        sheet.cell(column=11, row=sheet.max_row, value="Indicates anemia. This can be due to hematologic disorder, iron deficiency and bleeding.")

                        sheet.cell(column=11, row=sheet.max_row, value="therapy: in case of anemia Two 10 mg B12 pills a day for a month, in case of hematologic  disorder An injection of a hormone to encourage red blood cell production, in case of iron deficiency Two 10 mg pills of B12 a day for a month, in case of bleeding To be evacuated urgently to the hospital")


                file.save("Diagnose2.xlsx")
# *****************************************************
        def Iron(fh,iron,pregnant):
                file = openpyxl.load_workbook("Diagnose2.xlsx")
                sheet = file.active
                if (iron<60 and Gender()=="male" )or (iron<60*0.8 and Gender()=="female") or (pregnant==1 and Gender()=="female"):
                        sheet.cell(column=13, row=sheet.max_row, value="Usually indicates an inadequate diet or an increase in the need for iron or blood loss Following bleeding")
                        sheet.cell(column=13, row=sheet.max_row, value="therapy: in case of inadequate diet Schedule an appointment with a nutritionist, in case of Iron deficiency Two 10 mg pills of B12 a day for a month, in case of Bleeding To be evacuated urgently to the hospital")
                        fh.write(
                        '\nUsually indicates an inadequate diet or an increase in the need for iron or blood loss Following bleeding.')
                        fh.write(
                        '\ntherapy: in case of inadequate diet Schedule an appointment with a nutritionist, in case of Iron deficiency'
                        ' Two 10 mg pills of B12 a day for a month, in case of Bleeding To be evacuated urgently to the hospital')

                elif (iron>160 and Gender()=="male" )or (iron>160*0.8 and Gender()=="female"):
                        sheet.cell(column=13, row=sheet.max_row, value="May indicate iron poisoning.")
                        sheet.cell(column=13, row=sheet.max_row, value="therapy: to evacuate to the hospital")
                        fh.write('\nMay indicate iron poisoning.')
                        fh.write('\ntherapy: to evacuate to the hospital')

                else:
                        sheet.cell(column=13, row=sheet.max_row, value="\Its Ok")
                        fh.write('\nIts ok')
                file.save('Diagnose2.xlsx')
# *****************************************************
        def HDL(fh,hdl,Ethiopian):
                file = openpyxl.load_workbook("Diagnose2.xlsx")
                sheet = file.active
                if(hdl<29 and Gender()=="male")or(hdl<34 and Gender()=="female")or (hdl<29*1.2 and Gender()=="male" and Ethiopian==1) or (hdl<34*1.2 and Gender()=="female" and Ethiopian==1):
                        sheet.cell(column=14, row=sheet.max_row, value="\nMay indicate a risk of heart disease, hyperlipidemia or adult-onset diabetes.")
                        sheet.cell(column=14, row=sheet.max_row, value="\ntherapy: in case of heart disease to schedule an appointment with a nutritionist, in case of heart hyperlipidemia Schedule an appointment with a nutritionist, a 5 mg pill of Simobil daily for a week, in case of Adult diabetes Insulin adjustment for patient")
                        fh.write('\nMay indicate a risk of heart disease, hyperlipidemia or adult-onset diabetes.')
                        fh.write(
                        '\ntherapy: in case of heart disease to schedule an appointment with a nutritionist, in case of heart'
                        ' hyperlipidemia Schedule an appointment with a nutritionist, a 5 mg pill of Simobil daily for a week,'
                        ' in case of Adult diabetes Insulin adjustment for patient')


                elif(hdl>62 and Gender()=="male")or(hdl>82 and Gender()=="female")or (hdl>62*1.2 and Gender()=="male" and Ethiopian==1) or (hdl>82*1.2 and Gender()=="female" and Ethiopian==1):
                        sheet.cell(column=14, row=sheet.max_row, value="usually harmless.")
                        sheet.cell(column=14, row=sheet.max_row, value="therapy: Exercise raises good cholesterol levels.")
                        fh.write('\nusually harmless.')
                        fh.write('\ntherapy: Exercise raises good cholesterol levels.')

                else:
                        fh.write('\nIts ok')
                        sheet.cell(column=14, row=sheet.max_row, value="\Its Ok")
                file.save("Diagnose2.xlsx")
# *****************************************************


        def txtfile():

                name = NUserN.get()
                id=NIdN.get()
                age=Nl5N.get()
                fh = open('Diagnose.txt', 'w')

                file=openpyxl.load_workbook("Diagnose2.xlsx")
                sheet = file.active
                sheet.cell(column=1,row=sheet.max_row+1,value=name)
                sheet.cell(column=2,row=sheet.max_row, value=id)
                sheet.cell(column=3,row=sheet.max_row, value=age)

                if i.get()==1:
                        gen="Male"
                        sheet.cell(column=4,row=sheet.max_row,value="Male")
                else:
                        sheet.cell(column=4, row=sheet.max_row, value="Female")

                WbcCheck(int(Nl6N.get()),int(Nl5N.get()),fh)
                Neutcheck(fh, int(Nl7N.get()), int(Nl6N.get()))
                Lymph(fh, int(Nl8N.get()), int(Nl6N.get()))
                RBC(fh, int(Nl9N.get()), CheckVar1.get())
                HCT(fh, int(Nl10N.get()), CheckVar1.get())
                Urea(fh, int(Nl11N.get()), CheckVar3.get(), CheckVar2.get())
                Hb(fh, int(Nl12N.get()), int(Nl5N.get()))
                Creatinine(fh, int(Nl13N.get()), int(Nl5N.get()))
                Iron(fh, int(Nl14N.get()), CheckVar2.get())
                HDL(fh, int(Nl15N.get()), CheckVar4.get())
                Alkaline(fh, int(Nl16N.get()), CheckVar2.get(), CheckVar3.get())
                file.save("Diagnose2.xlsx")


                fh.write('name :' + name)
                fh.write('\nid :' + id)
                fh.write('\ngender :' + Gender())
                fh.write('\nage :' + age)
                fh.write('\n')
                WbcCheck(int(Nl6N.get()),int(Nl5N.get()),fh)
                fh.write('\n')
                Neutcheck(fh,int(Nl7N.get()),int(Nl6N.get()))
                fh.write('\n')
                Lymph(fh,int(Nl8N.get()),int(Nl6N.get()))
                fh.write('\n')
                RBC(fh,int(Nl9N.get()),CheckVar1.get())
                fh.write('\n')
                HCT(fh,int(Nl10N.get()),CheckVar1.get())
                fh.write('\n')
                Urea(fh,int(Nl11N.get()),CheckVar3.get(),CheckVar2.get())
                fh.write('\n')
                Hb(fh,int(Nl12N.get()),int(Nl5N.get()))
                fh.write('\n')
                Creatinine(fh,int(Nl13N.get()),int(Nl5N.get()))
                fh.write('\n')
                Iron(fh,int(Nl14N.get()),CheckVar2.get())
                fh.write('\n')
                HDL(fh,int(Nl15N.get()),CheckVar4.get())
                fh.write('\n')
                Alkaline(fh,int(Nl16N.get()),CheckVar2.get(),CheckVar3.get())
                fh.write('\n')
                fh.write('\n')
                fh.close()
                EndPage()




def secondPage():
        window.geometry("600x400")
        window.title("Doctor Profile")
        f2 = Frame()
        f2['bg'] = "#91b6bf"
        f2.place(x=0, y=0, width=600, height=400)
        image = Image.open('Doc.png')
        image = image.resize((100,100), Image.ANTIALIAS)
        img = ImageTk.PhotoImage(image)
        panel = Label(f2, image=img)
        panel.image=img
        panel.place(x=20,y=20)
        Newl1=Label(f2,text="Hello Doctor:  "+DocName(),fg="white",bg="#91b6bf",font=("times", 23,"bold"))
        Newl1.place(x=130,y=130)
        Newl2=Label(f2,text="To Adjust Patient's Profile:",fg="white",bg="#91b6bf",font=("times", 15))
        Newl2.place(x=50,y=230)
        Newl2 = Button(f2, text="Press For Adjusting", fg="white",bg="#91b6bf",command=thirdPage, font=("times", 15,"bold"))
        Newl2.place(x=230, y=280)

def DocName():
        return UserN.get()
def login():
        Nval = UserN.get()
        Pval = PassN.get()
        Idval = IdN.get()
        if Len3(Nval,Pval,Idval)==False:
            messagebox.showinfo("error","you must write all your details!!")
        elif RightPass(Pval)==False :
            messagebox.showinfo("password error","Check your password")
        elif RightUser(Nval)==False :
            messagebox.showinfo("User error","Check your user name")
        elif RightUserPass(Nval,Pval) == False:
            messagebox.showinfo("User or pass error", "Check your details")
        elif RightUser(Nval)==True and RightUser(Nval)==True and RightUserPass(Nval,Pval):
                secondPage()
f1=Frame()
f1['bg']="#E3E3E3"
f1.place(x=0,y=0,width=800,height=600)
l1=Label(f1,text="Please insert your information",fg="black",bg="#E3E3E3",font=("times", 23,"bold"))
l1.place(x=80,y=50)
l2=Label(f1,text="User Name :",fg="black",bg="#E3E3E3",font=("times", 15))
l2.place(x=110,y=130)
l3 = Label(f1, text="Password :", fg="black", bg="#E3E3E3",font=("times", 15))
l3.place(x=110, y=180)
l4 = Label(f1, text="Id :", fg="black", bg="#E3E3E3",font=("times", 15))
l4.place(x=110, y=230)
UserN=Entry(f1,borderwidth=3)
UserN.place(x=330,y=135)
PassN=Entry(f1,show = '*',borderwidth=3)
PassN.place(x=330,y=185)
IdN=Entry(f1,borderwidth=3)
IdN.place(x=330,y=235)
b1=Button(f1,text="Login", bg="#E3E3E3",fg="black",command=login,font=("times", 15,"bold"))
b1.place(x=265, y=300)

def Len3(name,pas,id):
        if len(name)!=0 and len(pas)!=0 and len(id)!=0  :
                return True
        return False

def RightUserPass(user,pas):
        if user == "maysaab" and pas == "#aaaaaa1":
                return True
        elif user == "adanto" and pas == "#aaaaaa1":
                return True
        return False

def RightPass(pas):
        count_special=0
        count_Num=0
        special_characters ='"!@#$%^&*()-+?_=,<>/"'
        for i in range(len(pas)):
                if pas[i] in special_characters:
                        count_special+=1
        for i in range(len(pas)):
                if pas[i].isdigit():
                        count_Num+=1
        if len(pas)>=8 and len(pas)<=10 and any(c.isalpha() for c in pas)==True and count_Num==1 and count_special==1:
                return True
        return False
def RightUser(user):
        count_Num=0
        for i in range(len(user)):
                if user[i].isdigit() == True:
                        count_Num +=1

        if len(user)>=6 and len(user)<=8 and count_Num<=2:
                return True
        return False

# *****************************************************
def WbcCheck(wbcValue, age, fh):
                file = openpyxl.load_workbook("Diagnose2.xlsx")
                sheet = file.active
                fh.write('\nWBC :')
                if age >= 18:
                        if wbcValue > 11000:
                                fh.write('\nMost often indicate the presence of an '
                                         ' infection,if there is a fever.In other cases,'
                                         ' very rare, may Very high values indicate blood disease or cancer.')
                                fh.write(
                                        '\ntherapy: antibiotic infection, in rare situations if there is a blood disease :'
                                        ' a combination of cyclophosphamide and corticosteroids if cancer : Entrectinib')
                                sheet.cell(column=5, row=sheet.max_row, value="Most often indicate the presence of an infection,if there is a fever.In other cases, very rare, may Very high values indicate blood disease or cancer.")
                                sheet.cell(column=5, row=sheet.max_row, value="\ntherapy: antibiotic infection, in rare situations if there is a blood disease : a combination of cyclophosphamide and corticosteroids if cancer : Entrectinib")

                        elif wbcValue < 4500:
                                fh.write(
                                        '\nIndicate viral disease, immune system failure and in very rare cases cancer.')
                                fh.write('\ntherapy: to rest at home if there was cancer: Entrectinib')
                                sheet.cell(column=5, row=sheet.max_row, value="\nIndicate viral disease, immune system failure and in very rare cases cancer.")
                                sheet.cell(column=5, row=sheet.max_row, value="\ntherapy: to rest at home if there was cancer: Entrectinib")

                        else:
                                fh.write('\nits ok')

                                sheet.cell(column=5,row=sheet.max_row,value="It's OK")


                elif age <= 17 and age >= 4:
                        if wbcValue > 15500:
                                fh.write('\nMost often indicate the presence of an '
                                         ' infection,if there is a fever.In other cases,'
                                         ' very rare, may Very high values indicate blood disease or cancer.')
                                fh.write(
                                        '\ntherapy: antibiotic infection, in rare situations if there is a blood disease :'
                                        ' a combination of cyclophosphamide and corticosteroids if cancer : Entrectinib')
                                sheet.cell(column=5, row=sheet.max_row, value="\nMost often indicate the presence of an infection,if there is a fever.In other cases, very rare, may Very high values indicate blood disease or cancer.")

                                sheet.cell(column=5, row=sheet.max_row, value="\ntherapy: antibiotic infection, in rare situations if there is a blood disease : a combination of cyclophosphamide and corticosteroids if cancer : Entrectinib")

                        elif wbcValue < 5500:
                                fh.write(
                                        '\nIndicate viral disease, immune system failure and in very rare cases cancer.')
                                fh.write('\ntherapy: to rest at home if there was cancer: Entrectinib')
                                sheet.cell(column=5, row=sheet.max_row, value="Indicate viral disease, immune system failure and in very rare cases cancer.")

                                sheet.cell(column=5, row=sheet.max_row, value="therapy: to rest at home if there was cancer: Entrectinib")

                        else:
                                fh.write('\nits ok')

                                sheet.cell(column=5, row=sheet.max_row, value="its ok")



                elif age >= 0 and age <= 3:
                        if wbcValue > 17500:
                                fh.write('\nMost often indicate the presence of an '
                                         ' infection,if there is a fever.In other cases,'
                                         ' very rare, may Very high values indicate blood disease or cancer.')
                                fh.write(
                                        '\ntherapy: antibiotic infection, in rare situations if there is a blood disease :'
                                        ' a combination of cyclophosphamide and corticosteroids if cancer : Entrectinib')
                                sheet.cell(column=5, row=sheet.max_row, value="\nMost often indicate the presence of an infection,if there is a fever.In other cases, very rare, may Very high values indicate blood disease or cancer.")

                                sheet.cell(column=5, row=sheet.max_row, value="\ntherapy: antibiotic infection, in rare situations if there is a blood disease : a combination of cyclophosphamide and corticosteroids if cancer : Entrectinib")

                        elif wbcValue < 6000:
                                fh.write(
                                        '\nIndicate viral disease, immune system failure and in very rare cases cancer.')
                                fh.write('\ntherapy: to rest at home if there was cancer: Entrectinib')
                                sheet.cell(column=5, row=sheet.max_row, value="\nIndicate viral disease, immune system failure and in very rare cases cancer.")

                                sheet.cell(column=5, row=sheet.max_row, value="\ntherapy: to rest at home if there was cancer: Entrectinib")

                        else:
                                fh.write('\nits ok')

                                sheet.cell(column=5, row=sheet.max_row, value="\nits ok")



                file.save("Diagnose2.xlsx")
# *****************************************************
def Neutcheck(fh,Neut,wbc):
        file = openpyxl.load_workbook("Diagnose2.xlsx")
        sheet = file.active
        fh.write('\nNeut :')

        if Neut > 0.28 *wbc and Neut<wbc*0.54:
                fh.write('\nIts ok')

                sheet.cell(column=6, row=sheet.max_row, value="\Its Ok")

        elif Neut<=wbc*0.28:
                fh.write(
                        '\nIndicate a disorder in the formation of blood, a tendency to bacterial infections and in rare cases - a process cancerous.')
                fh.write('\ntherapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month'
                         ' in case of bacterial infections antibiotic infection, if there was complications of cancer Entrectinib')
                sheet.cell(column=6, row=sheet.max_row, value="Indicate a disorder in the formation of blood, a tendency to bacterial infections and in rare cases - a process cancerous.")
                sheet.cell(column=6, row=sheet.max_row, value="therapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month in case of bacterial infections antibiotic infection, if there was complications of cancer Entrectinib")

        elif Neut>=wbc*0.54:
                fh.write('\nMost often indicate a bacterial infection.')
                fh.write('\ntherapy: in case of bacterial infections, antibiotic infection')
                sheet.cell(column=6, row=sheet.max_row, value="Most often indicate a bacterial infection.")
                sheet.cell(column=6, row=sheet.max_row, value="therapy: in case of bacterial infections, antibiotic infection")

        file.save("Diagnose2.xlsx")


# *****************************************************
def Lymph(fh,lymph,wbc):
        fh.write('\nLymph :')

        file = openpyxl.load_workbook("Diagnose2.xlsx")
        sheet = file.active
        if lymph>0.36*wbc and lymph<wbc*0.52:
                fh.write('\nIts ok')

                sheet.cell(column=7, row=sheet.max_row, value="\Its Ok")
        elif lymph<=wbc*0.36:
                fh.write('\nIndicate a problem in the formation of blood cells.')
                fh.write('\ntherapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month')
                sheet.cell(column=7, row=sheet.max_row, value="Indicate a problem in the formation of blood cells.")
                sheet.cell(column=7, row=sheet.max_row, value="therapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month")

        elif lymph>=wbc*0.52:
                fh.write('\nMay indicate a prolonged bacterial infection or lymphoma cancer.')
                fh.write(
                        '\ntherapy: in case of bacterial infections antibiotic infection.if there was complications of cancer Entrectinib')
                sheet.cell(column=7, row=sheet.max_row, value="May indicate a prolonged bacterial infection or lymphoma cancer.")
                sheet.cell(column=7, row=sheet.max_row, value="therapy: in case of bacterial infections antibiotic infection.if there was complications of cancer Entrectinib")
        file.save("Diagnose2.xlsx")
# *****************************************************
def RBC(fh,rbc,smoke):
        fh.write('\nRBC :')

        file = openpyxl.load_workbook("Diagnose2.xlsx")
        sheet = file.active
        if rbc<=6  and rbc>=4.5:
                fh.write('\nIts ok')

                sheet.cell(column=8, row=sheet.max_row, value="\Its Ok")
        elif rbc <4.5:
                fh.write('\nMay indicate anemia or severe bleeding.')
                fh.write('\ntherapy: in case of anemia Two 10 mg B12 pills a day for a month'
                         ', in case of bleeding To be evacuated urgently to the hospital')
                sheet.cell(column=8, row=sheet.max_row, value="May indicate anemia or severe bleeding.")
                sheet.cell(column=8, row=sheet.max_row, value="therapy: in case of anemia Two 10 mg B12 pills a day for a month , in case of bleeding To be evacuated urgently to the hospital")

        elif rbc >6:
                if smoke==1:
                        fh.write('\nMay indicate a disturbance in the blood production system.High levels were'
                                 ' also observed in smokers and in patients In lung diseases.')
                        fh.write(
                                '\ntherapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month'
                                ' Stop Smoking and Referral for X-ray of the lungs')
                        sheet.cell(column=8, row=sheet.max_row, value="May indicate a disturbance in the blood production system.High levels were also observed in smokers and in patients In lung diseases.")
                        sheet.cell(column=8, row=sheet.max_row, value="therapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month  Stop Smoking and Referral for X-ray of the lungs")


                else:
                        fh.write('\nMay indicate a disturbance in the blood production system.High levels were'
                                 ' also observed in smokers and in patients In lung diseases.')
                        fh.write(
                                '\ntherapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month'
                                ' Referral for X-ray of the lungs')
                        sheet.cell(column=8, row=sheet.max_row, value="May indicate a disturbance in the blood production system.High levels were also observed in smokers and in patients In lung diseases.")
                        sheet.cell(column=8, row=sheet.max_row, value="therapy: 10 mg pill of B12 a day for a month and 5 mg pill of folic acid a day for a month  Referral for X-ray of the lungs")

        file.save("Diagnose2.xlsx")
# *****************************************************
def Urea(fh,urea,asian,pregnant):
        file = openpyxl.load_workbook("Diagnose2.xlsx")
        sheet = file.active
        fh.write('\nUrea :')

        if urea<17 or (asian==1 and urea<17*1.1) or pregnant==1:
                fh.write('\nMalnutrition, low-protein diet or liver disease. It should be '
                         ' noted that during pregnancy the level of urination decreases.')
                fh.write('\ntherapy: in the Malnutrition case Schedule an appointment with a nutritionist'
                         ' in case of low protien diet Schedule an appointment with a nutritionist '
                         ' in case of Liver_disease Referral to a specific diagnosis for treatment')
                sheet.cell(column=10, row=sheet.max_row, value="Malnutrition, low-protein diet or liver disease. It should be noted that during pregnancy the level of urination decreases.")

                sheet.cell(column=10, row=sheet.max_row, value="therapy: in the Malnutrition case Schedule an appointment with a nutritionist in case of low protien diet Schedule an appointment with a nutritionist in case of Liver_disease Referral to a specific diagnosis for treatment")


        elif urea>43 or (asian==1 and urea>43*1.1):
                fh.write('\nMay indicate kidney disease, dehydration or a high protein diet.')
                fh.write('\ntherapy: in case of kidney disease balances blood sugar levels, in case of dehydration'
                         ' Complete rest while lying down, returning fluids to drinking in case of high protein'
                         ' diet Schedule an appointment with a nutritionist')
                sheet.cell(column=10, row=sheet.max_row, value="May indicate kidney disease, dehydration or a high protein diet.")

                sheet.cell(column=10, row=sheet.max_row, value="therapy: in case of kidney disease balances blood sugar levels, in case of dehydration Complete rest while lying down, returning fluids to drinking in case of high protein diet Schedule an appointment with a nutritionist")

        else:
                fh.write('\nIts ok')

                sheet.cell(column=10, row=sheet.max_row, value="\Its Ok")

        file.save("Diagnose2.xlsx")
# *****************************************************
def Creatinine(fh,creatinine,age):
        fh.write('\nCreatinine :')

        file = openpyxl.load_workbook("Diagnose2.xlsx")
        sheet = file.active
        if (age >= 0 and age <= 2 and creatinine < 0.2 )or (age >= 3 and age <= 17 and creatinine < 0.5 ) or (age >= 18 and age <= 59 and creatinine < 0.6) or (age>=60 and creatinine<0.6):
                fh.write(
                        '\nMost commonly seen in patients with very poor muscle mass and malnourished people who do not consume enough protein.')
                fh.write('\ntherapy: Schedule an appointment with a nutritionist')
                sheet.cell(column=12, row=sheet.max_row, value="Most commonly seen in patients with very poor muscle mass and malnourished people who do not consume enough protein.")
                sheet.cell(column=12, row=sheet.max_row, value="therapy: Schedule an appointment with a nutritionist")

        elif (age >=0 and age <=2 and creatinine>0.5 )or (age >=3 and age<=17 and creatinine>1 ) or (age >= 18 and age <= 59 and creatinine > 1) or (age>=60 and creatinine>1.2):
                fh.write(
                        '\nMay indicate a kidney problem and in severe cases kidney failure. High values are possible Also found in diarrhea and vomiting'
                        ' (causes of increased muscle breakdown and high levels of creatinine), muscle diseases And increased consumption of meat.')
                fh.write('\ntherapy: two pills 5 mg of Altman c3 turmeric a day for a month,'
                         ' to coordinate an appointment with a nutritionist')
                sheet.cell(column=12, row=sheet.max_row, value="May indicate a kidney problem and in severe cases kidney failure. High values are possible Also found in diarrhea and vomiting (causes of increased muscle breakdown and high levels of creatinine), muscle diseases And increased consumption of meat.")
                sheet.cell(column=12, row=sheet.max_row, value="therapy: two pills 5 mg of Altman c3 turmeric a day for a month, to coordinate an appointment with a nutritionist")

        else:
                fh.write('\nIts ok')

                sheet.cell(column=12, row=sheet.max_row, value="\Its Ok")
        file.save("Diagnose2.xlsx")
# *****************************************************
def Alkaline(fh,alkaline,pregnant,Asian):
        file = openpyxl.load_workbook("Diagnose2.xlsx")
        sheet = file.active
        fh.write('\nAlkaline :')

        if (alkaline<60 and Asian==1)or (alkaline<30 and Asian==0):
                fh.write(
                        '\nMay indicate a poor diet that lacks protein. Deficiency in vitamins like vitamin, vitamin B12, C, Vitamin B6 folic acid.')
                fh.write(
                        '\ntherapy: in case of poor diet Schedule an appointment with a nutritionist, in case of Vitamin Deficiency'
                        ' Referral for a blood test to identify the missing vitamins')
                sheet.cell(column=15, row=sheet.max_row, value="May indicate a poor diet that lacks protein. Deficiency in vitamins like vitamin, vitamin B12, C, Vitamin B6 folic acid.")
                sheet.cell(column=15, row=sheet.max_row, value="therapy: in case of poor diet Schedule an appointment with a nutritionist, in case of Vitamin Deficiency Referral for a blood test to identify the missing vitamins")

        elif (alkaline>120 and Asian==1)or (alkaline>90 and Asian==0)or pregnant==1:
                fh.write(
                        '\nMay indicate liver disease, biliary tract disease, pregnancy, hypothyroidism or Use of various medications.')
                fh.write(
                        '\ntherapy: in case of Liver disease Referral to a specific diagnosis for treatment, in case of tract disease Referral'
                        ' to surgical treatment in case of hypothyroidism Schedule an appointment with a nutritionist, a 5 mg pill of Simobil daily for a week')
                sheet.cell(column=15, row=sheet.max_row, value="May indicate liver disease, biliary tract disease, pregnancy, hypothyroidism or Use of various medications")
                sheet.cell(column=15, row=sheet.max_row, value="therapy: in case of Liver disease Referral to a specific diagnosis for treatment, in case of tract disease Referral to surgical treatment in case of hypothyroidism Schedule an appointment with a nutritionist, a 5 mg pill of Simobil daily for a wee")

        else:
                fh.write('\nIts ok')

                sheet.cell(column=15, row=sheet.max_row, value="\Its Ok")

        file.save("Diagnose2.xlsx")
window.mainloop()

